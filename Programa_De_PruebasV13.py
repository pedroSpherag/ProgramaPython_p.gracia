import paho.mqtt.client as paho
import win32com.client
from openpyxl import load_workbook, workbook
import time
import tkinter
from tkinter import ttk, messagebox
#-------------conectarse al mqtt -----------------------
def ConectarMqtt():
    broker="3.250.49.212"
    port=1883
    def on_publish(client,userdata,result):             #create function for callback
        #print("data published")
        pass
    def on_message(client, userdata, message):
        global imeix,datax,numDataEncontrados,mesgAlarmActivos,valorCont
        mensaje=str(message.payload.decode("utf-8"))        
        if message.topic=="Server/Logger":
            dataReci=DevuelveValor(mensaje,'"Data":"')
            imeiReci=DevuelveValor(mensaje,'"IMEI":"')
            valorCont=DevuelveValor(mensaje,'"Cont":"')
            if imeix==imeiReci and datax[numDataEncontrados]==dataReci:
                print("mesg recibido:"+imeix+", "+datax[numDataEncontrados])
                numDataEncontrados+=1              
            elif esImeiDeLista(imeiReci):
                if mesgAlarmActivos==True:
                    root = tkinter.Tk()
                    ttk.Frame(root).pack()
                    messagebox.showinfo("Se ha detectado 1 mensaje externo:", mensaje)
                    root.destroy()    
        else:
            print(imeix+": "+mensaje)
            numDataEncontrados+=1
             
    client1= paho.Client("Mosquitto Spherag")                           #create client object
    client1.username_pw_set("spheragBroker","Sphrg2020!!")
    client1.on_publish = on_publish                          #assign function to callback
    client1.on_message = on_message 
    client1.connect(broker,port, 28800)                                 #establish connection
    client1.loop_start() # Inicio del bucle
    client1.subscribe("Server/Logger")    
    return client1
#-----------metodo para cerrar ventana de mensajes
def cerrarMensajes():
    root.destroy()
#-----------se filtra el mensaje recibido    
def DevuelveValor(mensg,valor):#devuelve el valor del dato que se busca en el mensaje recibido
    ini=mensg.find(valor)   #se obtiene el valor de x en "Data":"x"
    datam=mensg[ini+8:ini+33]
    fin=datam.find('"')
    dato=datam[0:fin] 
    return dato 
#-------------te devuelve si el imei indicado es uno de los que se han leido del excel para trabajar
def esImeiDeLista(vImei):
    global imeis
    existe=False
    for i in range(len(imeis)):
        if imeis[i]==vImei:
            existe=True
    return existe
            
#------------abrir y leer fichero excel------------------------
def AbrirExcel(columna,hoja):
    wb=load_workbook("C:/Users/Pedro/Desktop/ImeiMqtt.xlsx")
    hoja=wb[hoja]
    x = []
    for row in hoja.iter_rows(): #Se leen todos los datos de la columna x            
        dato=row[columna-1].value
        datoStr=str(dato)                                   #Se traducen a string
        if dato != None and datoStr.isdigit() and len(datoStr)==15: # Unicamente se almacenan los datos que no sean un dato vacio, que contenga todo numeros y que sean 15 numeros.
            x.append(datoStr)                                          #Se almacenan los datos en una Array
    return  x
#-----------Obtener vector DataT---------------------------------------
def obtenerDataT(imeis):
    vectDataT=[]
    for x in range (len(imeis)):
        vectDataT.append(imeis[x] + "/DataT")
    return vectDataT
#-----------Obtener vector sombra Update---------------------------------------
def obtenerSombraUp(imeis):
    vectSombraUp=[]
    for x in range (len(imeis)):
        vectSombraUp.append("things/" + imeis[x] + "/shadow/update")
    return vectSombraUp
#-----------Obtener vector sombra Delete---------------------------------------
def obtenerSombraDel(imeis):
    vectSombraDel=[]
    for x in range (len(imeis)):
        vectSombraDel.append("things/" + imeis[x] + "/shadow/delete")
    return vectSombraDel
#-----------Obtener vector sombra Get---------------------------------------
def obtenerSombraGet(imeis):
    vectSombraGet=[]
    for x in range (len(imeis)):
        vectSombraGet.append("things/" + imeis[x] + "/shadow/get")
    return vectSombraGet

#------------guardar log en fichero excel------------------------
def GuardarLogExcel(imei,mensaje): 
    try:
        wb=load_workbook("C:/Users/Pedro/RIS IBERIA S.L/Spherag TEAM - General/16. Hardware&Firmware/MQTT/Codigo Python/Log.xlsx")
        hoja=wb['Log']
        existeImei=False
        fila=0
        mensajeMasFecha=mensaje+"_"+time.strftime("%d%m%Y_%H%M%S",time.localtime())
        for row in hoja.iter_rows(): 
            fila+=1
            if str(row[0].value) == imei:
                existeImei = True
                break
        if existeImei==False:
            hoja.cell(row=fila+1, column=1).value = imei
            hoja.cell(row=fila+1, column=2).value = mensajeMasFecha #se pone fila y columna a partir de 1 (primer dato (1,1))
            print("guardado en el log (Imei Nuevo)")
        else:
            columna=0
            existeHueco=False
            for column in hoja.iter_cols(): 
                columna+=1
                if column[fila-1].value == None:
                    existeHueco=True
                    hoja.cell(row=fila, column=columna).value = mensajeMasFecha
                    print("guardado en el log (Imei Existente)")
                    break
            if existeHueco==False:
                hoja.cell(row=fila, column=columna+1).value = mensajeMasFecha
                print("guardado en el log (Imei Existente, +)")
        wb.save('C:/Users/Pedro/RIS IBERIA S.L/Spherag TEAM - General/16. Hardware&Firmware/MQTT/Codigo Python/Log.xlsx')
    except:
        print("ERROR: No se ha podido abrir fichero LOG")
        time.sleep(1) 
        GuardarLogExcel(imei,mensaje)
# --------------prueba abrir y cerrar valvula----------
def ProbarSelenoide(client,imeis):
    global imeix
    imeisAveriados=[]
    eAutoManu = input("\nPulsa 'm' probar de forma manual [Open->Close], por defecto será automatica [Synchro] \n> ")
    print("Prueba selenoide:")
    dataT=obtenerDataT(imeis)
    for x in range (len(imeis)): 
        imeisAveriados.append(False)       
    for x in range (len(imeis)):
        imeix=imeis[x]
        print(str(x+1)+": "+imeis[x]+"....................") 
        print("Nota:Coloca el selenoide al siguiente dispositivo.")
        CparaContinuar() 
        if eAutoManu=="m":        
            for n in range (2):
                if n==0:
                    client.publish(dataT[x],'{"state":{"reported":{"M":"0","A":"1","GMT":"+01"}}}')  #publish (Se abren en todas las valvulas) 
                    imeisAveriados[x]=EsperaRespuesta("Programming","Manual: Valve open","")
                elif n==1:
                    client.publish(dataT[x],'{"state":{"reported":{"M":"0","A":"0","GMT":"+01"}}}')  #publish (Se abren en todas las valvulas) 
                    imeisAveriados[x]=EsperaRespuesta("Programming","Manual: Valve closed","")    
                time.sleep(1) 
                if imeisAveriados[x]:
                    break
        else:
            client.publish(dataT[x],'{"state":{"reported":{"GMT":"+01","Com":"IniSincro"}}}')  #publish (Se abren en todas las valvulas) 
            imeisAveriados[x]=EsperaRespuesta("Programming","Synchronized valve","")
            time.sleep(1)  
        PintaryGuardarResultado(imeis[x],"PruebaSelenoide",imeisAveriados[x])   
    ElimIMEISconError(imeis,imeisAveriados)   
# --------------manda una orden de abrir o cerrar valvula---------- 
def AbrCerrSelenoide(client,imeis):
    global imeix
    eAbrCerr = input("\nPulsa 'a' --> abrir valvula de forma manual, por defecto se mandará orden de cerrar \n> ")
    dataT=obtenerDataT(imeis)
    if eAbrCerr=='a':
        print("Abrir selenoide:")     
        for x in range (len(imeis)): 
            imeix=imeis[x]
            print(str(x+1)+": "+imeis[x]+"....................")       
            client.publish(dataT[x],'{"state":{"reported":{"M":"0","A":"1","GMT":"+01"}}} ')  #publish (Se abren en todas las valvulas) 
            errCom=EsperaRespuesta("Programming","Manual: Valve open","")
            PintaryGuardarResultado(imeis[x],"AbrirSelenoide",errCom)
            time.sleep(1)                 
    else:
        print("Cerrar selenoide:")     
        for x in range (len(imeis)): 
            imeix=imeis[x]
            print(str(x+1)+": "+imeis[x]+"....................") 
            client.publish(dataT[x],'{"state":{"reported":{"M":"0","A":"0","GMT":"+01"}}} ')  #publish (Se abren en todas las valvulas) 
            errCom=EsperaRespuesta("Programming","Manual: Valve closed","")
            PintaryGuardarResultado(imeis[x],"CerrarSelenoide",errCom)
            time.sleep(1) 
           
# --------------prueba contador-------- 
def ProbarContador(client,imeis):
    global valorCont,imeix
    contOk,imeisAveriados=[],[]
    dataT=obtenerDataT(imeis)
    for x in range (len(imeis)): 
        contOk.append(False)
        imeisAveriados.append(False)
    print("Prueba Contador:")
    for n in range (3):
        print("Parte "+str(n+1)+":")
        if n==0 or n==2 :
            for x in range (len(imeis)): 
                if imeisAveriados[x]==False:
                    imeix=imeis[x]
                    print(str(x+1)+": "+imeis[x]+"....................") 
                    client.publish(dataT[x],'{"state":{"reported":{"GMT":"+01","Com":"RESETCOUNTER"}}}')  #publish (Borra el contador) 
                    imeisAveriados[x]=EsperaRespuesta("Programming","","")
                    time.sleep(1) 
                    if imeisAveriados[x]==False:
                        client.publish(dataT[x],'{"state":{"reported":{}}}')  #publish (Actualiza los valores)
                        imeisAveriados[x]=EsperaRespuesta("Programming","","")
                        time.sleep(1)                  
        elif n==1 :
            algunOk=False
            for x in range (len(imeis)):
                if imeisAveriados[x]==False:
                    algunOk=True
                    break
            if algunOk:
                print("Lista de Imeis:")
                for x in range (len(imeis)):
                    if imeisAveriados[x]==False:
                        print (imeis[x])            
                print("\nNota:Junta 5 veces los cables de caudal de los dispositivos.")
                CparaContinuar()
                for x in range (len(imeis)): 
                    if imeisAveriados[x]==False:
                        imeix=imeis[x]
                        print(str(x+1)+": "+imeis[x]+"....................") 
                        client.publish(dataT[x],'{"state":{"reported":{}}}')  #publish (Actualiza los valores)
                        imeisAveriados[x]=EsperaRespuesta("Programming","","")
                        if int(valorCont)>1:
                            contOk[x]=True
                        time.sleep(1)              
    print("Resultado de las pruebas__")         
    for x in range (len(imeis)):
        if contOk[x] and imeisAveriados[x]==False:
            print(imeis[x]+ ": Ok test")  
            GuardarLogExcel(imeis[x],"ProbarContador(Ok)")
        else:
            print(imeis[x]+ ": Failed test") 
            GuardarLogExcel(imeis[x],"ProbarContador(Failed)")
    ElimIMEISconError(imeis,imeisAveriados)       
# --------------Borrar sombra------    
def BorraryPonerSombra(client,imeis,soloPonerSombra):
    global imeix
    print("Borrar y poner Sombra nueva:")
    print("Nota: En el caso de existir error de comunicación, se recomienda continuar hasta el final del programa. ")
    sombraUp=obtenerSombraUp(imeis)
    sombraDel=obtenerSombraDel(imeis)
    for x in range (len(imeis)): 
        imeix=imeis[x]
        print(str(x+1)+": "+imeis[x]+"....................")   
        #(borrar programas en dispositivo) 
        errCom=False
        if soloPonerSombra==False:
            client.publish(sombraUp[x],'{"state":{"reported":{"P0":"0","P1":"0","P2":"0","P3":"0","P4":"0","P5":"0","P6":"0","P7":"0","P8":"0","P9":"0","GMT":"+01"}}}')  
            errCom=EsperaRespuesta("Programming","","")
            time.sleep(1)
        client.publish(sombraDel[x],"")  #publish (Borrar Sombra) 
        time.sleep(1) 
        client.publish(sombraUp[x],'{"state":{"reported":{"M":"0","A":"0","EMode":"1","TPSM":"16","GMT":"+01"}}}')  #publish (se pone una sombra por defecto)
        time.sleep(1)
        PintaryGuardarResultado(imeis[x],"BorrarSombra",errCom)
#------------Poner en modo dormir y despertar cada x * 30 min    
def DormirXmin(client,imeis):
    tiempo = input("\nPulsa un nº de forma que [se despierta cada nº * 30min], por defecto será cada 30 min \n> ")
    if tiempo.isdigit() and int(tiempo)>=1 and int(tiempo)<=17:
        temp = 30 * int(tiempo)
    else:
        temp=30
        tiempo="1"
    print("Prueba selenoide:")
    funcionSM(client,imeis,"Dormir"+str(temp)+"m",'"EMode":"2","TPSM":"'+tiempo+'","GMT":"+01"')
#--------------Se obiente el valor de la sombra     
def ObtenerSombra(client,imeis,coment):
    global imeix
    print("Obtener Sombra "+coment+": ")
    sombraGet=obtenerSombraGet(imeis)
    dataT=obtenerDataT(imeis)
    for x in range (len(imeis)): 
        imeix=imeis[x]
        print(str(x+1)+": "+imeis[x]+"....................")             
        GuardarLogExcel(imeis[x],"ObtenerSombra"+coment)   
        client.subscribe(dataT[x])
        client.publish(sombraGet[x],"")  #publish (se obtiene valor sombra)
        EsperaRespuesta("","","")
        time.sleep(1)
        client.unsubscribe(dataT[x])
#--------------Comprobar con que velocidad contesta el dispositivo
def ComprobarVelocidad(client,imeis):
    global imeix
    print("Comprobar Velocidad de comunicacion:")  
    dataT=obtenerDataT(imeis)
    for x in range (len(imeis)): 
        imeix=imeis[x]
        print(str(x+1)+": "+imeis[x]+"....................")      
        tiempoAcumulado=0
        npruebas=3
        for i in range (npruebas):
            client.publish(dataT[x],'{"state":{"reported":{}}}')  #publicas y esperas respuesta
            start_time = time.time()
            errCom=EsperaRespuesta("Programming","","")
            if errCom == True:
                break            
            tiempoEnResponder = time.time() - start_time 
            tiempoAcumulado+=tiempoEnResponder
            print ("tiempo en responder, en el mensaje"+ str(i+1) +": "+ str(tiempoEnResponder) +" segundos" )   
            time.sleep(1)
        if errCom == False:
            tiempoMedio=tiempoAcumulado/npruebas
            print("tiempo medio en responder: "+ str(tiempoMedio) +" segundos")
            time.sleep(3)
        PintaryGuardarResultado(imeis[x],"ComprobarComunicacion",errCom)
#--------------Escuchar cualquier mensaje que reciba el dispositivo       
def escuchar(client,imeis):      
    global imeix
    print("Subscribirse y escuchar:")
    dataT=obtenerDataT(imeis)
    for x in range (len(imeis)): 
        imeix=imeis[x]
        print(str(x+1)+": "+imeis[x]+"....................")            
        GuardarLogExcel(imeis[x],"Subscribirse")  
        client.subscribe(dataT[x])
    opcion = input("Aprieta intro para dejar de escuchar > \n")
    for x in range (len(imeis)): 
        client.unsubscribe(dataT[x])
#-----------metodo para esperar a recibir respuesta del dispositivo despues de haber publicado           
def EsperaRespuesta(d1,d2,d3):
    global numDataEncontrados,datax
    datax[0],datax[1],datax[2]=d1,d2,d3
    i=0
    errorComunicacion=False
    numDataPrevistos=0
    if d2=="":
        numDataPrevistos=1
    elif d3=="":
        numDataPrevistos=2
    else:
        numDataPrevistos=3
    while numDataEncontrados<numDataPrevistos:
        i+=1
        time.sleep(0.5)
        if i==120:
            print("ERROR DE COMUNICACIÓN")
            letra = input("\nResto de teclas --> continuar esperando (Por defecto) \n'c'             --> saltar este imei y continuar con el siguiente \n> ")
            if letra=='c' and numDataEncontrados<numDataPrevistos:
                print("ok")
                errorComunicacion=True
                break                   
    numDataEncontrados=0
    for x in range(len(datax)):
        datax[x]=""
    return errorComunicacion
#------------Pulsa c para continuar
def CparaContinuar():
    print("Pulsa c para continuar..")
    while True: #Se espera a que se aprete la tecla c
        letra = input("> ")
        if letra=='c':
            print("ok")
            break 
#------------eliminar imeis con problemas de comunicación
def ElimIMEISconError(imeisLocal,imeisAveriados):
    global imeis
    imeis=[]
    for x in range (len(imeisLocal)):
        if imeisAveriados[x]==False:
            imeis.append(imeisLocal[x])
#------------Pinta en la pantalla y guarda en el log, el resultado de la prueba 
def PintaryGuardarResultado(imei,prueba,errCom):
    print("Resultado de la prueba__") 
    if errCom==False:
        print(imei+ ": Ok test")  
        GuardarLogExcel(imei,prueba+"(Ok)")
    else:
        print(imei+ ": Failed test") 
        GuardarLogExcel(imei,prueba+"(Failed)")     
#-----------------Parte inicial (se leen imeis y se muestran para su comprobación)
def ParteInicial(hoja,solicitudMensg):
    columnaExcel=1 #se lee todo los imeis que estan en la columna 1, que corresponde a la columna "A" de excel
    imeis = AbrirExcel(columnaExcel,hoja) 
    print("Lista de Imeis:")
    for x in range (len(imeis)):
        print (imeis[x])
    if (solicitudMensg):
        print ("\nNota: Comprueba que los imeis son correctos" )
    else:
        input("\nComprueba que los imeis son correctos y Pulsa intro > ")
    return imeis
#-----------------metodo general enviando a dataT
def funcionDT(client,imeis,orden,resp1,resp2,resp3):
    global imeix
    print(orden+": ") 
    dataT=obtenerDataT(imeis)
    for x in range (len(imeis)): 
        imeix=imeis[x]
        print(str(x+1)+": "+imeis[x]+"....................")    
        client.publish(dataT[x],'{"state":{"reported":{"GMT":"+01","Com":"'+orden+'"}}}')  #publish (Se abren en todas las valvulas) 
        errCom=EsperaRespuesta(resp1,resp2,resp3)
        PintaryGuardarResultado(imeis[x],orden,errCom)
        time.sleep(1)   
 # --------------metodo general modificando la sombra para modos de energia
def funcionSM(client,imeis,nombre,orden):
    global imeix
    print(nombre+": ")
    sombraUp=obtenerSombraUp(imeis)
    for x in range (len(imeis)): 
        imeix=imeis[x]
        print(str(x+1)+": "+imeis[x]+"....................")            
        GuardarLogExcel(imeis[x],nombre)  
        client.publish(sombraUp[x],'{"state":{"reported":{'+orden+'}}}')  #publish (Se cambia el modo de energía) 
        time.sleep(1)  
    

# ----------------------------------------COMIENZO PROGRAMA------------------------------------------------------------------------    
mesgAlarmActivos=True # De normal se reciben los mensajes de publicaciones "externas"
numDataEncontrados=0
datax=["","",""]
imeix=""
imeis=[] 
client=ConectarMqtt()
while True:  #menu de opciones
    print()
    print("[1]  prueba completa       \t[2]  probar selenoide       \t[3]  probar contador")
    print("[4]  modo Sleep(cada 8h)   \t[5]  modo Eco               \t[6]  modo Real time")
    print("[7]  modo Sleep(cada X min)  \t[8]  get Sombra             \t[9]  borrar y poner Sombra")
    print("[10] poner Sombra nueva    \t[11] C.Vel.Comunicacion     \t[12] escuchar mensajes")
    print("[13] Abrir/cerrar Manual   \t[14] Comandos especiales    \t[15] poner en hora")
    print("[0]  salir")
    opcion = input("> ")
    if opcion =="0":
        print("ok, salir...")
        break 
    
    #Aqui empiezan las opciones  -----------------------------------------------------------------------------------
    if opcion =="1":
        # PRUEBA COMPLETA-----------------
        mesgAlarmActivos=True
        imeis=ParteInicial("Imeis",False)
        #poner sombra
        BorraryPonerSombra(client,imeis,True)
        #prueba selenoide
        ProbarSelenoide(client,imeis)       
        #prueba contador
        ProbarContador(client,imeis) 
        #prueba dormir
        if (len(imeis))>0:
            DormirXmin(client,imeis)  
            print('NOTA: Se recomienda esperar 30 seg o comprobar que se ha dormido el ultimo dispositivo antes de cambiar de modo de energia (Real time, Eco, sleep).')   
        print("PROGRAMA FINALIZADO") 
        
    elif opcion =="2":
        # PRUEBA SELENOIDE
        mesgAlarmActivos=True
        imeis=ParteInicial("Imeis",True)
        ProbarSelenoide(client,imeis)
        print("PROGRAMA FINALIZADO")
        
    elif opcion =="3":
        # PRUEBA CONTADOR
        mesgAlarmActivos=True
        imeis=ParteInicial("Imeis",False)
        ProbarContador(client,imeis)
        print("PROGRAMA FINALIZADO") 
    
    elif opcion =="4":
        # MODO SLEEP(cada 8h)
        mesgAlarmActivos=False
        imeis=ParteInicial("Imeis",False)
        funcionSM(client,imeis,"Dormir8h",'"EMode":"2","TPSM":"16","GMT":"+01"')
        print("PROGRAMA FINALIZADO")        
        
    elif opcion =="5":
        # MODO ECO
        mesgAlarmActivos=False
        imeis=ParteInicial("Imeis",False)
        funcionSM(client,imeis,"ModoEco",'"EMode":"1","GMT":"+01"')
        print("PROGRAMA FINALIZADO")
        
    elif opcion =="6":
        # MODO REAL TIME
        mesgAlarmActivos=False
        imeis=ParteInicial("Imeis",False)
        funcionSM(client,imeis,"RealTime",'"EMode":"0","GMT":"+01"')
        print("PROGRAMA FINALIZADO")
        
    elif opcion =="7":
        # PRUEBA DORMIR(cada X min)
        mesgAlarmActivos=False
        imeis=ParteInicial("Imeis",False)
        DormirXmin(client,imeis)
        print("PROGRAMA FINALIZADO")
        
    elif opcion =="8":
        # OBTENER SOMBRA
        mesgAlarmActivos=False
        imeis=ParteInicial("Imeis",False)
        ObtenerSombra(client,imeis,"")
        print("PROGRAMA FINALIZADO")  
        
    elif opcion =="9":
        # BORRAR Y PONER SOMBRA
        mesgAlarmActivos=False
        imeis=ParteInicial("Imeis",False)        
        BorraryPonerSombra(client,imeis,False)
        print("PROGRAMA FINALIZADO") 
        
    elif opcion =="10":
        # PONER SOMBRA
        mesgAlarmActivos=False
        imeis=ParteInicial("Imeis",False)        
        BorraryPonerSombra(client,imeis,True)
        print("PROGRAMA FINALIZADO")         
          
    elif opcion =="11":
        #Comprobar Velocidad de comunicacion
        mesgAlarmActivos=True
        imeis=ParteInicial("Imeis",False)         
        ComprobarVelocidad(client,imeis)
        print("PROGRAMA FINALIZADO")
        
    elif opcion =="12":
        #Subscribirse para escuchar mensajes
        mesgAlarmActivos=True
        imeis=ParteInicial("Imeis",False) 
        escuchar(client,imeis)
        print("PROGRAMA FINALIZADO") 
        
    elif opcion =="13":
        #Abrir/Cerrar Manual
        mesgAlarmActivos=True
        imeis=ParteInicial("Imeis",True) 
        AbrCerrSelenoide(client,imeis)        
        print("PROGRAMA FINALIZADO") 
        
    elif opcion =="14":
        mesgAlarmActivos=True
        while True:  #menu de opciones
            print()
            print("[1]  UPCOUNTER     \t[2]  RESETCOUNTER    \t[3]  RESETUC")
            print("[4]  RESETMODEM    \t[5]  INFO            \t[6]  UPDATE")
            print('[0]  Menu Principal')
            sOpcion = input("> ")
            if sOpcion =="0":
                print("ok, volver menu principal...")
                break            
            elif sOpcion =="1":
                #UPCOUNTER
                imeis=ParteInicial("Imeis",False) 
                funcionDT(client,imeis,"UPCOUNTER","Programming","","")
                print("PROGRAMA FINALIZADO")
            elif sOpcion =="2":
                #RESETCOUNTER
                imeis=ParteInicial("Imeis",False) 
                funcionDT(client,imeis,"RESETCOUNTER","Programming","","")
                print("PROGRAMA FINALIZADO")
            elif sOpcion =="3":
                #RESETUC  
                imeis=ParteInicial("Imeis",False) 
                funcionDT(client,imeis,"RESETUC","Programming","","")
                print("PROGRAMA FINALIZADO")
            elif sOpcion =="4":
                #RESETMODEM
                imeis=ParteInicial("Imeis",False) 
                funcionDT(client,imeis,"RESETMODEM","Programming","","")
                print("PROGRAMA FINALIZADO")
            elif sOpcion =="5":
                #INFO
                imeis=ParteInicial("Imeis",False) 
                funcionDT(client,imeis,"INFO","Programming","Valve Info","")
                print("PROGRAMA FINALIZADO")
            elif sOpcion =="6":
                # UPDATE      
                imeis=ParteInicial("Imeis",False) 
                funcionDT(client,imeis,"UPDATE","Programming","Update","Programming")
                print("PROGRAMA FINALIZADO")
                
    elif opcion =="15":
        mesgAlarmActivos=False
        imeis=ParteInicial("Imeisph",False)   
        ObtenerSombra(client,imeis,"ph")
        print("PROGRAMA FINALIZADO")
        
client.loop_stop()
